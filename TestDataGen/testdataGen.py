from faker import Faker
from  openpyxl import Workbook

# fake = Faker()
#
# print(fake.name())
# print(fake.email())
# print(fake.address())
# print(fake.emoji())
# print(fake.credit_card_full())

wb = Workbook()
ws = wb.active

fake = Faker(['ja_JP','en_US'])

for i in range(1,10):
    for j in range(1,4):
        ws.cell(row=i, column=1).value= fake.name()
        ws.cell(row=i, column=2).value= fake.email()
        ws.cell(row=i, column=3).value= fake.address()
wb.save(filename="TestData.xlsx")

