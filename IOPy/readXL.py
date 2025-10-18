from openpyxl import load_workbook

wb = load_workbook(filename="DemoXl.xlsx")
sh = wb['Names']

row_count = sh.max_row
col_count = sh.max_column

print("Rows:", row_count)
print("Columns:", col_count)
print("---------------------")

for i in range(1, row_count + 1):
    for j in range(1, col_count + 1):
        print(sh.cell(row=i, column=j).value, end=' ')
    print()
