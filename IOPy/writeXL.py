from openpyxl import Workbook

wb = Workbook()
ws = wb.active

# ws['C3'] = "Raghu"
#
# test_date = [['Name','Age'],['Raghu',24],['Prakash',20],['Jaya', 50],['Ravi', 54]]
# for data in test_date:
#     ws.append(data)
#
# for i in range(1,6):
#     for j in range(1,4):
#         ws.cell(row = i, column=j).value = i+j

wb.save(filename="DemoXl.xlsx")


