import openpyxl

wb = openpyxl.load_workbook("Test.xlsx")
ws = wb.active
Dict = {}
#read value from row & column [matrix 2,2]
cell = ws.cell(row=2, column=1)
print(cell.value)

# write data in row & colum
val1 = ws.cell(row=2, column=2).value = "Raghuu"
print(val1)

print(ws.max_row)
print(ws.max_column)

print(ws['A5'].value)

# printing all data
for i in range(1, ws.max_row+1): # to get rows
    if ws.cell(row=i, column=2).value == "ewq": # check only Raghu in the column [till row 2] only single row / column
        for j in range(2, ws.max_column+1): # to get columns
            Dict[ws.cell(row=1, column=j).value] = ws.cell(row=2, column=j).value

print(Dict)

wb.close()