import time

import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait

path = "/Users/raghu/Downloads/download.xlsx"
driver = webdriver.Chrome()
driver.get("https://rahulshettyacademy.com/upload-download-test/")
driver.maximize_window()
driver.implicitly_wait(5)
Dict = {}
col_name = "price"
fruit_name = "Orange"
new_value = int(999)

#download file
driver.find_element(By.CSS_SELECTOR,"#downloadButton").click()
time.sleep(1)

def edit_value_excel(path, col_name, fruit_name, new_value):
    wb = openpyxl.load_workbook(path)
    ws = wb["Sheet1"]

    #find price in the file
    for i in range(1, ws.max_row + 1):
        if ws.cell(row=1, column=i).value == col_name:
            Dict["col"] = i

    for k in range(1,ws.max_row + 1):
        for j in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=j).value == fruit_name:
                Dict["row"] = k

    ws.cell(Dict["row"], Dict["col"]).value = new_value
    wb.save(path)

time.sleep(2)
# upload edited file
file_input = driver.find_element(By.CSS_SELECTOR,"#fileinput")
file_input.send_keys(path)
time.sleep(2)

wait = WebDriverWait(driver, 5)
locator = (By.CSS_SELECTOR, ".Toastify__toast-body")
wait.until(expected_conditions.visibility_of_element_located(locator))
print(driver.find_element(*locator).text)
time.sleep(3)

priceColumn = driver.find_element(By.XPATH,"//div[text()='Price']").get_attribute("data-column-id")
actual_price = driver.find_element(By.XPATH,"//div[text()='"+fruit_name+"']/parent::div/parent::div/div[@id='cell-"+priceColumn+"-undefined']").text
assert int(actual_price) == new_value
time.sleep(1)


driver.quit()